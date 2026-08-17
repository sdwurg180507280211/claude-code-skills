#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import plistlib
import re
import sys
import uuid
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - surfaced as a friendly CLI error
    load_workbook = None

NAME_CANDIDATES = [
    "current_name",
    "original_name",
    "公众号名称",
    "快捷方式名称",
    "名称",
    "name",
    "label",
]
URL_CANDIDATES = [
    "target_url",
    "URL",
    "url",
    "链接",
    "文章URL",
    "主页URL",
]
ICON_CANDIDATES = [
    "icon_path",
    "图标路径",
    "头像路径",
    "icon",
]
PNG_SIGNATURE = b"\x89PNG\r\n\x1a\n"


@dataclass(frozen=True)
class WebClipEntry:
    name: str
    url: str
    icon_path: str = ""


def normalize_header(value: object) -> str:
    return re.sub(r"[\s_\-]+", "", str(value or "").strip()).lower()


def choose_column(headers: list[str], explicit: str | None, candidates: list[str], required: bool) -> str | None:
    by_normalized = {normalize_header(h): h for h in headers if h}
    if explicit:
        key = normalize_header(explicit)
        if key not in by_normalized:
            raise ValueError(f"找不到列：{explicit}")
        return by_normalized[key]
    for candidate in candidates:
        key = normalize_header(candidate)
        if key in by_normalized:
            return by_normalized[key]
    if required:
        raise ValueError(f"无法自动识别必需列；可选列名：{', '.join(candidates)}")
    return None


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        return [{str(k or "").strip(): str(v or "").strip() for k, v in row.items()} for row in reader]


def read_xlsx_rows(path: Path, sheet: str | None) -> list[dict[str, str]]:
    if load_workbook is None:
        raise RuntimeError("缺少 openpyxl；请先执行 pip install -r requirements.txt")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    values = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(values)
    except StopIteration:
        return []
    headers = [str(v or "").strip() for v in raw_headers]
    rows: list[dict[str, str]] = []
    for raw in values:
        row = {headers[i]: str(raw[i] or "").strip() for i in range(min(len(headers), len(raw))) if headers[i]}
        if any(row.values()):
            rows.append(row)
    return rows


def read_rows(path: Path, sheet: str | None) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path, sheet)
    raise ValueError("输入仅支持 .csv / .xlsx / .xlsm")


def is_http_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def resolve_icon_path(raw: str, input_path: Path) -> Path | None:
    if not raw:
        return None
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = input_path.parent / path
    return path.resolve()


def load_png(path: Path) -> bytes:
    data = path.read_bytes()
    if not data.startswith(PNG_SIGNATURE):
        raise ValueError(f"图标必须是 PNG：{path}")
    if len(data) > 1024 * 1024:
        raise ValueError(f"图标文件超过 1 MB：{path}")
    return data


def collect_entries(
    input_path: Path,
    sheet: str | None,
    name_column: str | None,
    url_column: str | None,
    icon_column: str | None,
) -> tuple[list[WebClipEntry], list[dict]]:
    rows = read_rows(input_path, sheet)
    if not rows:
        return [], []
    headers = list(rows[0].keys())
    name_col = choose_column(headers, name_column, NAME_CANDIDATES, required=True)
    url_col = choose_column(headers, url_column, URL_CANDIDATES, required=True)
    icon_col = choose_column(headers, icon_column, ICON_CANDIDATES, required=False)

    entries: list[WebClipEntry] = []
    skipped: list[dict] = []
    seen_names: dict[str, str] = {}
    for index, row in enumerate(rows, start=2):
        name = str(row.get(name_col or "", "") or "").strip()
        url = str(row.get(url_col or "", "") or "").strip()
        icon = str(row.get(icon_col or "", "") or "").strip() if icon_col else ""
        if not name:
            skipped.append({"row": index, "reason": "missing_name"})
            continue
        if not is_http_url(url):
            skipped.append({"row": index, "name": name, "reason": "missing_or_invalid_url", "url": url})
            continue
        if name in seen_names:
            skipped.append(
                {
                    "row": index,
                    "name": name,
                    "reason": "duplicate_name",
                    "kept_url": seen_names[name],
                    "skipped_url": url,
                }
            )
            continue
        seen_names[name] = url
        entries.append(WebClipEntry(name=name, url=url, icon_path=icon))
    return entries, skipped


def stable_uuid(seed: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_URL, seed)).upper()


def payload_suffix(name: str, url: str) -> str:
    return hashlib.sha256(f"{name}\n{url}".encode("utf-8")).hexdigest()[:16]


def build_webclip_payload(
    entry: WebClipEntry,
    profile_identifier: str,
    input_path: Path,
    fullscreen: bool,
    removable: bool,
) -> dict:
    suffix = payload_suffix(entry.name, entry.url)
    payload_id = f"{profile_identifier}.webclip.{suffix}"
    payload = {
        "FullScreen": fullscreen,
        "IgnoreManifestScope": False,
        "IsRemovable": removable,
        "Label": entry.name,
        "Precomposed": True,
        "URL": entry.url,
        "PayloadDisplayName": entry.name,
        "PayloadIdentifier": payload_id,
        "PayloadType": "com.apple.webClip.managed",
        "PayloadUUID": stable_uuid(payload_id),
        "PayloadVersion": 1,
    }
    icon_path = resolve_icon_path(entry.icon_path, input_path)
    if icon_path:
        if not icon_path.is_file():
            raise ValueError(f"图标文件不存在：{icon_path}")
        payload["Icon"] = load_png(icon_path)
    return payload


def build_profile(
    entries: list[WebClipEntry],
    input_path: Path,
    profile_name: str,
    profile_identifier: str,
    organization: str,
    fullscreen: bool,
    removable: bool,
) -> dict:
    payloads = [
        build_webclip_payload(
            entry,
            profile_identifier=profile_identifier,
            input_path=input_path,
            fullscreen=fullscreen,
            removable=removable,
        )
        for entry in entries
    ]
    return {
        "PayloadContent": payloads,
        "PayloadDisplayName": profile_name,
        "PayloadDescription": f"为 {len(entries)} 个微信公众账号/文章创建 iOS 主屏幕 Web Clip。",
        "PayloadIdentifier": profile_identifier,
        "PayloadOrganization": organization,
        "PayloadType": "Configuration",
        "PayloadUUID": stable_uuid(profile_identifier),
        "PayloadVersion": 1,
    }


def write_summary(path: Path, entries: list[WebClipEntry], skipped: list[dict], output_path: Path) -> None:
    summary = {
        "generated": len(entries),
        "skipped": len(skipped),
        "output": str(output_path),
        "entries": [{"name": item.name, "url": item.url, "icon_path": item.icon_path} for item in entries],
        "skipped_rows": skipped,
    }
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="批量生成 iOS 微信公众号 Web Clip .mobileconfig")
    parser.add_argument("--input", required=True, help="输入 .csv / .xlsx；可直接使用 wechat-account-bookmarks 的 wechat_accounts.csv")
    parser.add_argument("--sheet", default=None, help="Excel Sheet；默认第一个")
    parser.add_argument("--name-column", default=None, help="名称列；默认自动识别 current_name/original_name/公众号名称/快捷方式名称")
    parser.add_argument("--url-column", default=None, help="URL 列；默认优先识别 target_url")
    parser.add_argument("--icon-column", default=None, help="可选 PNG 图标路径列；相对路径以输入文件目录为基准")
    parser.add_argument("--output", default="wechat-ios-webclips.mobileconfig", help="输出 .mobileconfig")
    parser.add_argument("--profile-name", default="微信公众号快捷方式", help="iOS 配置描述文件显示名称")
    parser.add_argument("--profile-id", default="com.my-skills.wechat-ios-shortcuts", help="配置描述文件 PayloadIdentifier")
    parser.add_argument("--organization", default="my-skills", help="配置描述文件组织名称")
    parser.add_argument("--fullscreen", action="store_true", help="以全屏 Web App 形式打开；默认使用普通浏览器 UI")
    parser.add_argument("--non-removable", action="store_true", help="将 Web Clip 设为不可单独删除；默认可删除")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    if not input_path.is_file():
        print(f"输入文件不存在：{input_path}", file=sys.stderr)
        return 2
    if output_path.suffix.lower() != ".mobileconfig":
        print("输出文件扩展名应为 .mobileconfig", file=sys.stderr)
        return 2

    try:
        entries, skipped = collect_entries(
            input_path,
            sheet=args.sheet,
            name_column=args.name_column,
            url_column=args.url_column,
            icon_column=args.icon_column,
        )
        if not entries:
            print("没有可生成的 Web Clip；请检查名称列和 URL 列。", file=sys.stderr)
            return 3
        profile = build_profile(
            entries,
            input_path=input_path,
            profile_name=args.profile_name,
            profile_identifier=args.profile_id,
            organization=args.organization,
            fullscreen=args.fullscreen,
            removable=not args.non_removable,
        )
    except (OSError, RuntimeError, ValueError, KeyError) as exc:
        print(str(exc), file=sys.stderr)
        return 2

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(plistlib.dumps(profile, fmt=plistlib.FMT_XML, sort_keys=False))
    summary_path = output_path.with_suffix(".summary.json")
    write_summary(summary_path, entries, skipped, output_path)
    print(f"Generated {len(entries)} Web Clips: {output_path}")
    if skipped:
        print(f"Skipped {len(skipped)} rows; details: {summary_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
