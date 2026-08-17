#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path


NAME_CANDIDATES = ["快捷方式名称", "公众号名称", "内容名称", "名称", "name"]
FOLDER_CANDIDATES = ["文件夹结构", "分类", "文件夹", "folder", "category"]
URL_CANDIDATES = ["URL", "url", "链接", "公众号链接", "文章链接", "历史链接", "fallback_article_url"]
BIZ_CANDIDATES = ["__biz", "biz", "account_biz", "公众号biz"]
TARGET_CANDIDATES = ["目标类型", "target_type", "书签目标", "preferred_target", "target"]


def normalize_target_pref(value: object) -> str:
    """Normalize a per-row target preference to auto/homepage/article."""
    v = str(value or "").strip().lower()
    if v in {"article", "文章", "文章书签"}:
        return "article"
    if v in {"homepage", "home", "主页", "主页书签"}:
        return "homepage"
    return "auto"


@dataclass(frozen=True)
class InputEntry:
    name: str
    folder: str = ""
    url: str = ""
    biz: str = ""
    target_pref: str = "auto"

    def identity_key(self) -> str:
        payload_data = {"name": self.name, "url": self.url, "biz": self.biz}
        if self.target_pref != "auto":
            payload_data["target_pref"] = self.target_pref
        payload = json.dumps(
            payload_data,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _norm_header(value: object) -> str:
    return str(value or "").strip()


def choose_column(
    headers: list[str],
    requested: str | None,
    candidates: list[str],
    required: bool,
) -> str | None:
    if requested:
        if requested not in headers:
            raise ValueError(f"找不到列：{requested}；现有列：{headers}")
        return requested
    for candidate in candidates:
        if candidate in headers:
            return candidate
    if required:
        raise ValueError(f"无法自动识别名称列；现有列：{headers}")
    return None


def _entry_from_mapping(
    row: dict,
    name_col: str,
    folder_col: str | None,
    url_col: str | None,
    biz_col: str | None,
    target_col: str | None = None,
) -> InputEntry | None:
    name = str(row.get(name_col, "") or "").strip()
    if not name:
        return None
    folder = str(row.get(folder_col, "") or "").strip() if folder_col else ""
    url = str(row.get(url_col, "") or "").strip() if url_col else ""
    biz = str(row.get(biz_col, "") or "").strip() if biz_col else ""
    target_pref = normalize_target_pref(row.get(target_col)) if target_col else "auto"
    return InputEntry(name=name, folder=folder, url=url, biz=biz, target_pref=target_pref)


def load_csv(
    path: Path,
    name_column: str | None,
    folder_column: str | None,
    url_column: str | None,
    biz_column: str | None,
    target_column: str | None = None,
) -> tuple[list[InputEntry], dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        headers = [_norm_header(h) for h in (reader.fieldnames or [])]
        name_col = choose_column(headers, name_column, NAME_CANDIDATES, True)
        folder_col = choose_column(headers, folder_column, FOLDER_CANDIDATES, False)
        url_col = choose_column(headers, url_column, URL_CANDIDATES, False)
        biz_col = choose_column(headers, biz_column, BIZ_CANDIDATES, False)
        target_col = choose_column(headers, target_column, TARGET_CANDIDATES, False)
        entries: list[InputEntry] = []
        for row in reader:
            entry = _entry_from_mapping(row, name_col, folder_col, url_col, biz_col, target_col)
            if entry:
                entries.append(entry)
    return entries, {
        "sheet": None,
        "name_column": name_col,
        "folder_column": folder_col,
        "url_column": url_col,
        "biz_column": biz_col,
        "target_column": target_col,
    }


def load_xlsx(
    path: Path,
    sheet_name: str | None,
    name_column: str | None,
    folder_column: str | None,
    url_column: str | None,
    biz_column: str | None,
    target_column: str | None = None,
) -> tuple[list[InputEntry], dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 .xlsx 需要 openpyxl，请先 pip install -r requirements.txt") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"找不到 Sheet：{sheet_name}；现有：{wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], {
            "sheet": ws.title,
            "name_column": name_column or "",
            "folder_column": folder_column,
            "url_column": url_column,
            "biz_column": biz_column,
            "target_column": target_column or "",
        }

    headers = [_norm_header(x) for x in header_row]
    name_col = choose_column(headers, name_column, NAME_CANDIDATES, True)
    folder_col = choose_column(headers, folder_column, FOLDER_CANDIDATES, False)
    url_col = choose_column(headers, url_column, URL_CANDIDATES, False)
    biz_col = choose_column(headers, biz_column, BIZ_CANDIDATES, False)
    target_col = choose_column(headers, target_column, TARGET_CANDIDATES, False)

    indexes = {
        "name": headers.index(name_col),
        "folder": headers.index(folder_col) if folder_col else None,
        "url": headers.index(url_col) if url_col else None,
        "biz": headers.index(biz_col) if biz_col else None,
        "target": headers.index(target_col) if target_col else None,
    }

    entries: list[InputEntry] = []
    for row in rows:
        if indexes["name"] >= len(row):
            continue
        values: dict[str, object] = {name_col: row[indexes["name"]]}
        for key, col in (("folder", folder_col), ("url", url_col), ("biz", biz_col)):
            idx = indexes[key]
            if col and idx is not None and idx < len(row):
                values[col] = row[idx]
        if target_col and indexes["target"] is not None and indexes["target"] < len(row):
            values[target_col] = row[indexes["target"]]
        entry = _entry_from_mapping(values, name_col, folder_col, url_col, biz_col, target_col)
        if entry:
            entries.append(entry)

    return entries, {
        "sheet": ws.title,
        "name_column": name_col,
        "folder_column": folder_col,
        "url_column": url_col,
        "biz_column": biz_col,
        "target_column": target_col,
    }


def load_entries(
    path: Path,
    sheet_name: str | None = None,
    name_column: str | None = None,
    folder_column: str | None = None,
    url_column: str | None = None,
    biz_column: str | None = None,
    target_column: str | None = None,
) -> tuple[list[InputEntry], dict]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv(path, name_column, folder_column, url_column, biz_column, target_column)
    if suffix == ".xlsx":
        return load_xlsx(
            path,
            sheet_name,
            name_column,
            folder_column,
            url_column,
            biz_column,
            target_column,
        )
    raise ValueError("只支持 .xlsx 和 .csv 输入")


def identity_fingerprint(entries: list[InputEntry]) -> str:
    """Fingerprint identity inputs only; folder edits do not invalidate cached identity results."""
    canonical = sorted({entry.identity_key() for entry in entries})
    payload = "\n".join(canonical)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def normalize_folder(folder: str, strip_prefix: str | None) -> list[str]:
    if not folder:
        return []
    parts = [part.strip() for part in re.split(r"\s*>\s*", folder) if part.strip()]
    if len(parts) <= 1 and "/" in folder:
        parts = [part.strip() for part in folder.split("/") if part.strip()]
    if strip_prefix and parts and parts[0] == strip_prefix:
        parts = parts[1:]
    return parts


def atomic_write_json(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def load_json(path: Path, default):
    if not path.is_file():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return default
